from aiogram import Dispatcher, types
from app.handlers.common import AdressList
from app.handlers.adresses import check_address
from aiogram.dispatcher import FSMContext
import os
from aiogram.dispatcher.filters import Text
import openpyxl
import sqlite3
from aiogram.dispatcher.filters import Filter


class IsAdmin(Filter):
    key = 'is_admin'

    async def check(self, message: types.Message):
        with sqlite3.connect('app/users.db') as conn:
            cursor = conn.cursor()
            id = message.chat.id
            select_stmt = f"select role from users_table where user_chat_id={id}"
            cursor.execute(select_stmt)
            result = cursor.fetchone()
            if result[0] == 'admin':
                return True
            else:
                return False


def format_file(format_of_file):
    formats = ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel',
               'text/csv']
    for i in formats:
        if i == format_of_file:
            return False
    return True


async def excel_parcing(message: types.Message, state: FSMContext):
    typer = message.document.mime_type

    if format_file(typer):
        await message.answer(f'Файл не является файлом формата xls, xlsx или csv.')
        return

    direc = os.getcwd()
    last_file = await message.document.download(destination_dir=f'{direc}/app/files/{message.chat.id}')
    excel_file = openpyxl.load_workbook(last_file.name)
    sheet = excel_file['Лист1']
    maxrow = sheet.max_row
    user_data = await state.get_data()
    adresses = user_data['adress']

    for x in range(1, maxrow + 1):
        value_cell = sheet.cell(row=x, column=1).value
        if value_cell is not None and 'Челябинск' in value_cell:
            coords = check_address(value_cell[25:])
            new_adress = [value_cell[25:], ' ✅', coords]
            adresses.append(new_adress)

    await state.update_data(adress=adresses)

    keyboard = types.InlineKeyboardMarkup()

    for i in adresses:
        index_element = adresses.index(i)
        print(index_element)
        str_index_element = str(index_element)
        button1 = types.InlineKeyboardButton(text=i[0] + i[1], callback_data='delete' + str_index_element)
        keyboard.add(button1)
    await message.answer("Адресс был успешно добавлен введите следующий "
                         "адресс, либо постройте маршрут", reply_markup=keyboard)


async def delete_file(message: types.Message):
    direc = os.getcwd()
    directory = f'{direc}\\app\\documents\\'
    for i in os.walk(directory):
        print(i[2])
        os.remove(f'{directory}\\{i[2][0]}')
    await message.answer('worked')


def register_handlers_excel(dp: Dispatcher):
    dp.register_message_handler(excel_parcing, IsAdmin(), content_types=['document'],
                                state=AdressList.waiting_for_add_adress)
    dp.register_message_handler(delete_file, Text(equals='Удалить'), state=AdressList.waiting_for_add_adress)
