import openpyxl
import sqlite3
import csv


async def xlsx_reader(last_file):
    excel_file = openpyxl.load_workbook(last_file.name)
    sheet = excel_file['Лист1']
    value_a1 = sheet.cell(row=1, column=1).value
    if value_a1 != 'Номера телефонов':
        text = 'Файл не соотвествует формату, указнному выше'
        return text
    maxrow = sheet.max_row
    answer_text = 'Файл загружен. '
    answer_text_for_error = 'Отчет:'
    print(maxrow)
    for x in range(2, maxrow + 1):
        value_cell = sheet.cell(row=x, column=1).value
        if value_cell is None:
            continue
        symbols = '-+ '
        value = str(value_cell).translate(str.maketrans('', '', symbols))
        if value[0] == '8':
            value = value.lstrip('8')
        with sqlite3.connect('app/users.db') as conn:
            cursor = conn.cursor()
            sql = f"select username from users_table where telephone_number like '%{value}'"
            cursor.execute(sql)
            result = cursor.fetchone()
            if result is None:
                answer_text_for_error = answer_text_for_error + f'\n<b><u>{value_cell}</u></b> - Номер не ' \
                                                                f'зарегистрирован или ' \
                                                                f'некорректно указан в файле '
            else:
                sql_update = f"update users_table set role='admin' where telephone_number like '%{value}'"
                cursor.execute(sql_update)
                answer_text_for_error = answer_text_for_error + f'\n<b><u>{value_cell}</u></b> - Роль пользователя ' \
                                                                f'успешно заменена '
    return answer_text + answer_text_for_error


async def csv_reader(last_file):
    answer_text = 'Файл загружен. '
    answer_text_for_error = 'Отчет:'
    with open(last_file) as file:
        reader = csv.reader(file)
        list_of_csv = list(reader)
        if list_of_csv[0][0] != 'Номера телефонов':
            text = 'Файл не соотвествует формату, указнному выше'
            return text
        range_of_csv = list_of_csv[1:]
        print(range_of_csv)
        for i in range_of_csv:
            symbols = '-+ '
            value = str(i[0]).translate(str.maketrans('', '', symbols))
            if value[0] == '8':
                value = value.lstrip('8')
            print(value)
            with sqlite3.connect('app/users.db') as conn:
                cursor = conn.cursor()
                sql = f"select username from users_table where telephone_number like '%{value}'"
                cursor.execute(sql)
                result = cursor.fetchone()
                if result is None:
                    answer_text_for_error = answer_text_for_error + f'\n<b><u>{i[0]}</u></b> - Номер не ' \
                                                                    f'зарегистрирован или ' \
                                                                    f'некорректно указан в файле '
                else:
                    sql_update = f"update users_table set role='admin' where telephone_number like '%{value}'"
                    cursor.execute(sql_update)
                    answer_text_for_error = answer_text_for_error + f'\n<b><u>{i[0]}</u></b> - Роль пользователя ' \
                                                                    f'успешно заменена '
    return answer_text + answer_text_for_error
